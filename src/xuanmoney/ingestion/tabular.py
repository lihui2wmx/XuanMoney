from __future__ import annotations

import csv
from collections.abc import Iterable, Mapping
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from xuanmoney.domain import DimensionalRow, IncomeStatement
from xuanmoney.semantic.registry import (
    SemanticMappingError,
    resolve_dimensional_columns,
    resolve_income_statement_columns,
)


class TabularIngestionError(ValueError):
    pass


def _decimal(value: Any, *, field: str, row_number: int) -> Decimal:
    if isinstance(value, bool):
        raise TabularIngestionError(
            f"row {row_number}: boolean is not a valid numeric value for {field!r}"
        )
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if value is None:
        raise TabularIngestionError(f"row {row_number}: missing value for {field!r}")

    text = str(value).strip().replace(",", "")
    if not text:
        raise TabularIngestionError(f"row {row_number}: missing value for {field!r}")
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise TabularIngestionError(
            f"row {row_number}: invalid decimal {value!r} for {field!r}"
        ) from exc


def _required_text(
    row: Mapping[str, Any],
    column: str,
    *,
    field: str,
    row_number: int,
) -> str:
    value = row.get(column)
    if value is None or not str(value).strip():
        raise TabularIngestionError(f"row {row_number}: {field} is required")
    return str(value).strip()


def income_statements_from_rows(
    rows: Iterable[Mapping[str, Any]],
    *,
    source: str = "tabular",
) -> list[IncomeStatement]:
    materialized = list(rows)
    if not materialized:
        return []

    try:
        columns = resolve_income_statement_columns(materialized[0].keys())
    except SemanticMappingError as exc:
        raise TabularIngestionError(str(exc)) from exc

    optional_money_fields = (
        "operating_expenses",
        "other_income",
        "other_expenses",
        "taxes",
    )
    statements: list[IncomeStatement] = []
    for row_number, row in enumerate(materialized, start=2):
        payload: dict[str, Any] = {
            "period": _required_text(
                row,
                columns["period"],
                field="period",
                row_number=row_number,
            ),
            "source": f"{source}:row:{row_number}",
            "revenue": _decimal(
                row.get(columns["revenue"]), field="revenue", row_number=row_number
            ),
            "cogs": _decimal(row.get(columns["cogs"]), field="cogs", row_number=row_number),
        }

        if "currency" in columns:
            currency = row.get(columns["currency"])
            if currency is not None and str(currency).strip():
                payload["currency"] = str(currency).strip()

        for field in optional_money_fields:
            source_column = columns.get(field)
            if source_column is None:
                continue
            raw_value = row.get(source_column)
            if raw_value is None or not str(raw_value).strip():
                continue
            payload[field] = _decimal(raw_value, field=field, row_number=row_number)

        statements.append(IncomeStatement.model_validate(payload))

    return statements


def dimensional_rows_from_rows(
    rows: Iterable[Mapping[str, Any]],
    *,
    source: str = "tabular",
) -> list[DimensionalRow]:
    """Normalize explicit one-dimensional business rows.

    Required semantics are `period`, `dimension`, `member`, `revenue`, and `cogs`.
    Unknown columns are ignored; business dimensions are never inferred from arbitrary
    source column names.
    """

    materialized = list(rows)
    if not materialized:
        return []

    try:
        columns = resolve_dimensional_columns(materialized[0].keys())
    except SemanticMappingError as exc:
        raise TabularIngestionError(str(exc)) from exc

    normalized: list[DimensionalRow] = []
    for row_number, row in enumerate(materialized, start=2):
        payload: dict[str, Any] = {
            "period": _required_text(
                row,
                columns["period"],
                field="period",
                row_number=row_number,
            ),
            "dimension": _required_text(
                row,
                columns["dimension"],
                field="dimension",
                row_number=row_number,
            ),
            "member": _required_text(
                row,
                columns["member"],
                field="member",
                row_number=row_number,
            ),
            "source": f"{source}:row:{row_number}",
            "revenue": _decimal(
                row.get(columns["revenue"]), field="revenue", row_number=row_number
            ),
            "cogs": _decimal(row.get(columns["cogs"]), field="cogs", row_number=row_number),
        }

        if "currency" in columns:
            currency = row.get(columns["currency"])
            if currency is not None and str(currency).strip():
                payload["currency"] = str(currency).strip()

        normalized.append(DimensionalRow.model_validate(payload))

    return normalized


def _load_csv(path: Path) -> list[IncomeStatement]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise TabularIngestionError("CSV file has no header row")
        rows = list(reader)
    return income_statements_from_rows(rows, source=path.name)


def _load_dimensional_csv(path: Path) -> list[DimensionalRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise TabularIngestionError("CSV file has no header row")
        rows = list(reader)
    return dimensional_rows_from_rows(rows, source=path.name)


def _xlsx_rows(
    path: Path,
    *,
    sheet_name: str | None,
) -> tuple[list[dict[str, Any]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is part of normal install
        raise TabularIngestionError("openpyxl is required for .xlsx ingestion") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            worksheet = workbook[workbook.sheetnames[0]]
        else:
            if sheet_name not in workbook.sheetnames:
                raise TabularIngestionError(f"worksheet {sheet_name!r} does not exist")
            worksheet = workbook[sheet_name]

        iterator = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration as exc:
            raise TabularIngestionError("Excel worksheet is empty") from exc

        headers = [
            str(value).strip() if value is not None else f"__unnamed_{index}"
            for index, value in enumerate(raw_headers, start=1)
        ]
        rows = [dict(zip(headers, values, strict=False)) for values in iterator]
        resolved_sheet_name = worksheet.title
    finally:
        workbook.close()

    return rows, resolved_sheet_name


def _load_xlsx(path: Path, *, sheet_name: str | None) -> list[IncomeStatement]:
    rows, resolved_sheet_name = _xlsx_rows(path, sheet_name=sheet_name)
    return income_statements_from_rows(
        rows,
        source=f"{path.name}:{resolved_sheet_name}",
    )


def _load_dimensional_xlsx(
    path: Path,
    *,
    sheet_name: str | None,
) -> list[DimensionalRow]:
    rows, resolved_sheet_name = _xlsx_rows(path, sheet_name=sheet_name)
    return dimensional_rows_from_rows(
        rows,
        source=f"{path.name}:{resolved_sheet_name}",
    )


def load_income_statements(
    path: str | Path,
    *,
    sheet_name: str | None = None,
) -> list[IncomeStatement]:
    """Load normalized v0.1 income statements from CSV or XLSX.

    The loader performs only explicit semantic alias resolution. It does not ask an LLM
    to infer unknown columns and fails closed when required finance fields are missing.
    """

    file_path = Path(path)
    suffix = file_path.suffix.casefold()
    if suffix == ".csv":
        return _load_csv(file_path)
    if suffix == ".xlsx":
        return _load_xlsx(file_path, sheet_name=sheet_name)
    raise TabularIngestionError(
        f"unsupported tabular format {file_path.suffix!r}; expected .csv or .xlsx"
    )


def load_dimensional_rows(
    path: str | Path,
    *,
    sheet_name: str | None = None,
) -> list[DimensionalRow]:
    """Load canonical one-dimensional business rows from CSV or XLSX."""

    file_path = Path(path)
    suffix = file_path.suffix.casefold()
    if suffix == ".csv":
        return _load_dimensional_csv(file_path)
    if suffix == ".xlsx":
        return _load_dimensional_xlsx(file_path, sheet_name=sheet_name)
    raise TabularIngestionError(
        f"unsupported tabular format {file_path.suffix!r}; expected .csv or .xlsx"
    )
